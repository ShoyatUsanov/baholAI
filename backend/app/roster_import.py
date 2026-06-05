"""Parse an uploaded student roster (.xlsx / .xls / .csv) into clean rows.

Tolerant: matches Uzbek or English headers, falls back to positional columns
(name, username, password, level) when there's no recognizable header row.
"""
from __future__ import annotations

import csv
import io

from openpyxl import load_workbook

HEADER_ALIASES = {
    "name": {"name", "ism", "ismi", "fio", "f.i.o", "fish", "to'liq ism", "toliq ism", "full name"},
    "username": {"username", "login", "user", "user name"},
    "password": {"password", "parol", "pass", "parol(login)"},
    "level": {"level", "daraja", "kurs", "bosqich", "guruh"},
}


def _norm(v) -> str:
    return str(v).strip() if v is not None else ""


def _read_grid(data: bytes, filename: str) -> list[list[str]]:
    name = (filename or "").lower()
    if name.endswith(".csv") or name.endswith(".txt"):
        text = data.decode("utf-8-sig", errors="replace")
        # accept comma OR semicolon separated
        sep = ";" if text.count(";") > text.count(",") else ","
        rows = list(csv.reader(io.StringIO(text), delimiter=sep))
    else:
        wb = load_workbook(io.BytesIO(data), read_only=True, data_only=True)
        ws = wb.active
        rows = [list(r) for r in ws.iter_rows(values_only=True)]
    grid = [[_norm(c) for c in r] for r in rows]
    return [r for r in grid if any(r)]


def parse_roster(data: bytes, filename: str) -> list[dict]:
    """Returns [{name, username, password, level}]. Skips rows lacking name/username."""
    grid = _read_grid(data, filename)
    if not grid:
        return []

    header = [c.lower() for c in grid[0]]
    colmap: dict[str, int] = {}
    for idx, h in enumerate(header):
        for key, aliases in HEADER_ALIASES.items():
            if h in aliases and key not in colmap:
                colmap[key] = idx
    if "name" in colmap and "username" in colmap:
        body = grid[1:]
    else:
        colmap = {"name": 0, "username": 1, "password": 2, "level": 3}
        body = grid

    out: list[dict] = []
    for r in body:
        def cell(key: str) -> str:
            i = colmap.get(key)
            return r[i] if i is not None and i < len(r) else ""

        name, username = cell("name"), cell("username")
        if not name or not username:
            continue
        out.append({
            "name": name,
            "username": username,
            "password": cell("password") or "student123",
            "level": cell("level") or None,
        })
    return out
